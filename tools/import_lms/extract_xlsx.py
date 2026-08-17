"""Чтение выгрузки ServiceGuru: уроки (+встроенные картинки) и тесты.

Два неочевидных момента формата, выясненных разведкой выгрузки:

1. Картинки лежат в xlsx как embedded media, а якоря drawing'ов ДРЕЙФУЮТ
   относительно строк типа «Изображение» (совпадение точное лишь на 76 листах
   из 115). Поэтому сопоставление строго ПОРЯДКОВОЕ: N-й якорь листа ↔ N-я
   строка-картинка. Лишние строки без якоря пропускаются (16 таких в выгрузке —
   в основном обложки уроков).
2. Правильный ответ теста помечен ЖИРНЫМ шрифтом ячейки (`font.bold`), другого
   признака нет. Читать обязательно с `read_only=False`, иначе стилей нет.
"""

from __future__ import annotations

import hashlib
import re
import zipfile
from dataclasses import dataclass, field
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl

_NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
}

IMAGE_ROW_TYPES = {"Изображение", "Изображение урока"}


@dataclass
class LessonRow:
    kind: str  # Глава | Текст | Изображение | Изображение урока | Видео | Загруженное видео
    content: str
    caption: str
    image_md5: str | None = None  # для строк-картинок, если якорь нашёлся


@dataclass
class LessonSheet:
    title: str  # имя листа (может быть обрезано до 28 символов)
    rows: list[LessonRow]


@dataclass
class Question:
    prompt: str
    options: list[str]
    correct: list[int]
    image_md5: str | None = None

    @property
    def qtype(self) -> str:
        return "multi" if len(self.correct) > 1 else "single"


@dataclass
class QuizSheet:
    title: str
    questions: list[Question] = field(default_factory=list)


def _sheet_order(path: Path) -> list[str]:
    """Имена листов в порядке xl/worksheets/sheetN.xml (для маппинга drawings)."""
    with zipfile.ZipFile(path) as zf:
        wb = ET.fromstring(zf.read("xl/workbook.xml"))
        rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    target_by_id = {r.get("Id"): r.get("Target") for r in rels.findall("rel:Relationship", _NS)}
    out = []
    for sheet in wb.find("main:sheets", _NS):
        rid = sheet.get(f"{{{_NS['r']}}}id")
        target = (target_by_id.get(rid) or "").split("/")[-1]
        out.append((sheet.get("name"), target))
    return out


def _anchors_by_sheet(path: Path) -> dict[str, list[tuple[int, bytes]]]:
    """{имя листа: [(0-based строка якоря, байты картинки), …]} в порядке якорей.

    Уроки сопоставляются ПОРЯДКОВО (строки дрейфуют), тесты — по номеру
    строки (там якорь стоит ровно на строке вопроса).
    """
    result: dict[str, list[tuple[int, bytes]]] = {}
    with zipfile.ZipFile(path) as zf:
        names = set(zf.namelist())
        for sheet_name, target in _sheet_order(path):
            result[sheet_name] = []
            rels_path = f"xl/worksheets/_rels/{target}.rels"
            if rels_path not in names:
                continue
            sheet_rels = ET.fromstring(zf.read(rels_path))
            drawing_targets = [
                r.get("Target")
                for r in sheet_rels.findall("rel:Relationship", _NS)
                if r.get("Type", "").endswith("/drawing")
            ]
            for dt in drawing_targets:
                drawing_name = "xl/" + dt.replace("../", "")
                if drawing_name not in names:
                    continue
                drawing = ET.fromstring(zf.read(drawing_name))
                drels_path = (
                    f"xl/drawings/_rels/{drawing_name.split('/')[-1]}.rels"
                )
                if drels_path not in names:
                    continue
                drels = ET.fromstring(zf.read(drels_path))
                media_by_id = {
                    r.get("Id"): "xl/" + r.get("Target").replace("../", "")
                    for r in drels.findall("rel:Relationship", _NS)
                }
                for anchor in drawing:
                    blip = anchor.find(".//{http://schemas.openxmlformats.org/drawingml/2006/main}blip")
                    if blip is None:
                        continue
                    embed = blip.get(f"{{{_NS['r']}}}embed")
                    media = media_by_id.get(embed)
                    if not (media and media in names):
                        continue
                    frm = anchor.find("xdr:from", _NS)
                    row_el = frm.find("xdr:row", _NS) if frm is not None else None
                    row_idx = int(row_el.text) if row_el is not None and row_el.text else -1
                    result[sheet_name].append((row_idx, zf.read(media)))
    return result


def _md5(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()  # noqa: S324 — дедуп файлов, не крипто


def read_lessons(path: Path, media_sink: dict[str, bytes]) -> list[LessonSheet]:
    """Листы-уроки файла `<slug>_lessons_export.xlsx`.

    Найденные картинки складываются в `media_sink` (md5 → байты) и
    привязываются к строкам порядково.
    """
    anchors = _anchors_by_sheet(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: list[LessonSheet] = []
    for ws in wb.worksheets:
        images = [data for _row, data in anchors.get(ws.title, [])]
        img_idx = 0
        rows: list[LessonRow] = []
        for raw in ws.iter_rows(values_only=True):
            if not raw or not raw[0]:
                continue
            kind = str(raw[0]).strip()
            if kind == "Тип":  # шапка
                continue
            content = str(raw[1]).strip() if len(raw) > 1 and raw[1] else ""
            caption = str(raw[2]).strip() if len(raw) > 2 and raw[2] else ""
            row = LessonRow(kind=kind, content=content, caption=caption)
            # Строка-картинка без своего якоря остаётся без файла (16 таких
            # в выгрузке) — при сборке урока она просто пропускается.
            if kind in IMAGE_ROW_TYPES and img_idx < len(images):
                data = images[img_idx]
                img_idx += 1
                digest = _md5(data)
                media_sink.setdefault(digest, data)
                row.image_md5 = digest
            rows.append(row)
        sheets.append(LessonSheet(title=ws.title.strip(), rows=rows))
    wb.close()
    return sheets


def _correct_indices(cells) -> tuple[list[str], list[int]]:
    """Варианты ответа + индексы правильных (жирные ячейки).

    Пустые ячейки отбрасываются ДО вычисления индексов — иначе дырка между
    вариантами сдвинула бы `correct`.
    """
    options: list[str] = []
    correct: list[int] = []
    for cell in cells:
        value = cell.value
        if value is None or not str(value).strip():
            continue
        text = str(value).strip()
        if text in options:  # дубль варианта — не плодим
            continue
        if cell.font is not None and cell.font.bold:
            correct.append(len(options))
        options.append(text)
    return options, correct


def read_quizzes(path: Path, media_sink: dict[str, bytes]) -> list[QuizSheet]:
    """Листы-тесты файла `<slug>_surveys_export.xlsx` (bold = правильный)."""
    anchors = _anchors_by_sheet(path)
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    quizzes: list[QuizSheet] = []
    for ws in wb.worksheets:
        # Иллюстрация вопроса: якорь стоит РОВНО на строке вопроса (проверено
        # на выгрузке) — привязка по номеру строки, а не по порядку.
        image_by_row = {row: data for row, data in anchors.get(ws.title, []) if row >= 0}
        header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
        try:
            prompt_col = header.index("Вопрос")
        except ValueError:
            continue  # лист без вопросов («Empty name»)
        answer_cols = [i for i, name in enumerate(header) if name == "Ответ"]
        quiz = QuizSheet(title=ws.title.strip())
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
            if prompt_col >= len(row) or not row[prompt_col].value:
                continue
            prompt = str(row[prompt_col].value).strip()
            options, correct = _correct_indices([row[i] for i in answer_cols if i < len(row)])
            if not options:
                continue
            image_md5 = None
            data = image_by_row.get(row_idx)
            if data is not None:
                image_md5 = _md5(data)
                media_sink.setdefault(image_md5, data)
            quiz.questions.append(
                Question(prompt=prompt, options=options, correct=correct, image_md5=image_md5)
            )
        quizzes.append(quiz)
    wb.close()
    return quizzes


_MENU_HTML = re.compile(r"<[^>]+>")


def read_menu(path: Path, media_sink: dict[str, bytes] | None = None) -> list[dict]:
    """Лист «Меню» экспорта ассортимента.

    Фото — миниатюры 150×100 из ServiceGuru (1–3 КБ): для витрины мелковато,
    но лучше пустой плитки (решение пользователя 2026-08-17). Якорь стоит на
    строке позиции — привязка по номеру строки, как в тестах.
    """
    anchors = _anchors_by_sheet(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    image_by_row = {row: data for row, data in anchors.get(ws.title, []) if row >= 0}
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    out: list[dict] = []
    for row_idx, raw in enumerate(rows[1:], start=1):
        if not raw or not raw[idx["название"]]:
            continue

        def cell(name: str, row=raw) -> str:
            i = idx.get(name)
            if i is None or i >= len(row) or row[i] is None:
                return ""
            return str(row[i]).strip()

        photo_md5 = None
        data = image_by_row.get(row_idx)
        if data is not None and media_sink is not None:
            photo_md5 = _md5(data)
            media_sink.setdefault(photo_md5, data)

        out.append(
            {
                "title": cell("название"),
                "category": cell("категория"),
                "path": cell("путь к подкатегории"),
                "description_html": cell("описание"),
                "composition": cell("ингредиенты"),
                "photo_md5": photo_md5,
            }
        )
    wb.close()
    return out
