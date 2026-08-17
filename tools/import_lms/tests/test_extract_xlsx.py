"""Извлечение уроков/тестов из xlsx-выгрузки.

Фикстуры синтетические (сама выгрузка `LMS/` в репозиторий не коммитится),
но повторяют формат ServiceGuru: колонки, bold-пометку правильного ответа,
встроенные картинки с якорями.
"""

from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Font

from tools.import_lms.extract_xlsx import read_lessons, read_menu, read_quizzes

_DRAWING_TMPL = """<?xml version="1.0" encoding="UTF-8"?>
<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
          xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">{anchors}</xdr:wsDr>"""

_ANCHOR_TMPL = """<xdr:oneCellAnchor>
  <xdr:from><xdr:col>{col}</xdr:col><xdr:colOff>0</xdr:colOff>
            <xdr:row>{row}</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
  <xdr:ext cx="900000" cy="900000"/>
  <xdr:pic><xdr:nvPicPr><xdr:cNvPr id="{n}" name="image{n}.jpg"/><xdr:cNvPicPr/></xdr:nvPicPr>
    <xdr:blipFill><a:blip xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"
                          r:embed="rId{n}"/><a:stretch/></xdr:blipFill>
    <xdr:spPr/></xdr:pic><xdr:clientData/></xdr:oneCellAnchor>"""

_DRAWING_RELS_TMPL = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{rels}</Relationships>"""


def _attach_images(path: Path, images: list[tuple[int, int, bytes]]) -> None:
    """Добавить в xlsx встроенные картинки с якорями (col, row, данные).

    Пишем OOXML руками: openpyxl для этого требует Pillow, а нам важно ещё и
    проверить собственный разбор drawings/rels на настоящей структуре пакета.
    """
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path) as src, zipfile.ZipFile(tmp, "w") as dst:
        for item in src.infolist():
            data = src.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                data = data.replace(
                    b"</worksheet>",
                    b'<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/'
                    b'2006/relationships" r:id="rIdDraw"/></worksheet>',
                )
            dst.writestr(item, data)
        anchors = "".join(
            _ANCHOR_TMPL.format(col=col, row=row, n=i + 1)
            for i, (col, row, _data) in enumerate(images)
        )
        dst.writestr("xl/drawings/drawing1.xml", _DRAWING_TMPL.format(anchors=anchors))
        rels = "".join(
            f'<Relationship Id="rId{i + 1}" Type="http://schemas.openxmlformats.org/'
            f'officeDocument/2006/relationships/image" Target="../media/image{i + 1}.jpg"/>'
            for i in range(len(images))
        )
        dst.writestr("xl/drawings/_rels/drawing1.xml.rels", _DRAWING_RELS_TMPL.format(rels=rels))
        dst.writestr(
            "xl/worksheets/_rels/sheet1.xml.rels",
            _DRAWING_RELS_TMPL.format(
                rels='<Relationship Id="rIdDraw" Type="http://schemas.openxmlformats.org/'
                'officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>'
            ),
        )
        for i, (_col, _row, data) in enumerate(images):
            dst.writestr(f"xl/media/image{i + 1}.jpg", data)
    shutil.move(tmp, path)


@pytest.fixture
def lessons_file(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Урок про кофе"
    ws.append(["Тип", "Контент", "Описание"])
    ws.append(["Изображение урока", "", ""])
    ws.append(["Глава", "Первая глава", ""])
    ws.append(["Текст", "<p>Абзац</p>", ""])
    ws.append(["Изображение", "", "Подпись к схеме"])
    ws.append(["Видео", "https://youtu.be/abc123", ""])
    path = tmp_path / "x_lessons_export.xlsx"
    wb.save(path)
    # Якоря стоят на «чужих» строках (6 и 8) — как дрейф в реальной выгрузке;
    # привязка обязана идти по порядку, а не по номеру строки.
    _attach_images(path, [(1, 6, b"COVER-BYTES"), (1, 8, b"SCHEME-BYTES")])
    return path


@pytest.fixture
def quiz_file(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Для закрепления"
    ws.append(["Изображение", "Время, с", "Вопрос", "Ответ", "Ответ", "Ответ", "Комментарий"])
    ws.append([None, 90, "Один правильный?", "нет", "да", "тоже нет", ""])
    ws.cell(row=2, column=5).font = Font(bold=True)
    ws.append([None, 60, "Два правильных?", "верно", "неверно", "тоже верно", ""])
    ws.cell(row=3, column=4).font = Font(bold=True)
    ws.cell(row=3, column=6).font = Font(bold=True)
    # Пустая ячейка МЕЖДУ вариантами не должна сдвинуть индексы правильных.
    ws.append([None, 60, "С дыркой?", "первый", None, "третий", ""])
    ws.cell(row=4, column=6).font = Font(bold=True)
    path = tmp_path / "x_surveys_export.xlsx"
    wb.save(path)
    # Иллюстрация вопроса: колонка A, строка ровно первого вопроса (0-based 1).
    _attach_images(path, [(0, 1, b"QUESTION-IMAGE")])
    return path


def test_lesson_rows_and_ordinal_image_binding(lessons_file: Path):
    media: dict[str, bytes] = {}
    sheets = read_lessons(lessons_file, media)
    assert len(sheets) == 1
    rows = sheets[0].rows
    assert [r.kind for r in rows] == [
        "Изображение урока",
        "Глава",
        "Текст",
        "Изображение",
        "Видео",
    ]
    covers = [r for r in rows if r.kind == "Изображение урока"]
    figures = [r for r in rows if r.kind == "Изображение"]
    # Якоря стояли на «чужих» строках — привязка порядковая, а не по строке.
    assert covers[0].image_md5 is not None
    assert figures[0].image_md5 is not None
    assert covers[0].image_md5 != figures[0].image_md5
    assert len(media) == 2
    assert figures[0].caption == "Подпись к схеме"
    assert rows[-1].content == "https://youtu.be/abc123"


def test_lesson_image_row_without_anchor_stays_unbound(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Тип", "Контент", "Описание"])
    ws.append(["Изображение урока", "", ""])  # обложки нет в файле (16 таких)
    ws.append(["Текст", "<p>Есть только текст</p>", ""])
    path = tmp_path / "y_lessons_export.xlsx"
    wb.save(path)
    media: dict[str, bytes] = {}
    rows = read_lessons(path, media)[0].rows
    assert rows[0].image_md5 is None
    assert media == {}


def test_quiz_bold_marks_correct_answers(quiz_file: Path):
    media: dict[str, bytes] = {}
    quizzes = read_quizzes(quiz_file, media)
    assert len(quizzes) == 1
    q1, q2, q3 = quizzes[0].questions
    assert (q1.options, q1.correct, q1.qtype) == (["нет", "да", "тоже нет"], [1], "single")
    assert (q2.correct, q2.qtype) == ([0, 2], "multi")
    # Пустой вариант отброшен ДО вычисления индексов: «третий» стал вторым.
    assert (q3.options, q3.correct) == (["первый", "третий"], [1])


def test_quiz_image_bound_by_row(quiz_file: Path):
    media: dict[str, bytes] = {}
    questions = read_quizzes(quiz_file, media)[0].questions
    assert questions[0].image_md5 is not None
    assert questions[1].image_md5 is None
    assert len(media) == 1


def test_quiz_sheet_without_answers_skipped(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empty name"
    ws.append(["Изображение", "Время, с", "Вопрос", "Комментарий"])
    path = tmp_path / "z_surveys_export.xlsx"
    wb.save(path)
    quizzes = read_quizzes(path, {})
    assert quizzes[0].questions == []


def test_duplicate_option_text_deduped(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Изображение", "Время, с", "Вопрос", "Ответ", "Ответ", "Ответ"])
    ws.append([None, 60, "Дубли?", "да", "да", "нет"])
    ws.cell(row=2, column=4).font = Font(bold=True)
    path = tmp_path / "d_surveys_export.xlsx"
    wb.save(path)
    q = read_quizzes(path, {})[0].questions[0]
    assert q.options == ["да", "нет"]
    assert q.correct == [0]


def test_read_menu(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Меню"
    ws.append(
        ["id", "категория", "путь к подкатегории", "тип", "название", "описание", "ингредиенты"]
    )
    ws.append(
        [85, "Горячие напитки", "Напитки/Кофе", "наименование", "Раф", "<p>Вкусно</p>", "Кофе"]
    )
    path = tmp_path / "menu.xlsx"
    wb.save(path)
    _attach_images(path, [(6, 1, b"MENU-THUMB")])  # колонка «изображение», строка позиции
    photos: dict[str, bytes] = {}
    items = read_menu(path, photos)
    assert len(items) == 1
    item = items[0]
    assert item["title"] == "Раф"
    assert item["category"] == "Горячие напитки"
    assert item["path"] == "Напитки/Кофе"
    assert item["description_html"] == "<p>Вкусно</p>"
    assert item["composition"] == "Кофе"
    # Миниатюра позиции привязана по номеру строки и сложена в media_sink.
    assert item["photo_md5"] is not None
    assert list(photos.values()) == [b"MENU-THUMB"]


def test_read_menu_without_sink_skips_photos(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Меню"
    ws.append(["id", "категория", "путь к подкатегории", "тип", "название"])
    ws.append([1, "Супы", "Супы", "наименование", "Борщ"])
    path = tmp_path / "menu2.xlsx"
    wb.save(path)
    items = read_menu(path)
    assert items[0]["photo_md5"] is None
